"""
Step 6b: Fetch data from Wind MCP.
Saves fetched data to a staging file (JSON) — does NOT modify the database.

When Wind MCP is available, this script calls the MCP tool for each series.
When Wind MCP is NOT available, it can simulate a fetch from Excel for pipeline testing.
"""
import sys, json
from pathlib import Path
from datetime import datetime

sys.path.insert(0, str(Path(__file__).resolve().parent))
from lib import get_db
from transforms import transform_with_audit as apply_transforms_with_audit


def fetch_via_wind_mcp(plan_entry):
    """
    Fetch data via Wind MCP for a single series (production path).

    Uses wind_mcp_adapter.wind_mcp_query → economic_data.get_economic_data CLI.
    Maps the plan's indicator + last DB dates into a Wind MCP query, returns
    only NEW observations (after last_date) for downstream overlap validation.

    Returns: {date: value} of fetched data (incl. overlap dates), or None.
    """
    try:
        from wind_mcp_adapter import wind_mcp_query
    except ImportError:
        return None

    indicator = plan_entry.get("query") or plan_entry.get("wind_indicator")
    if not indicator:
        return None

    # Fetch window: from second-to-last DB date through today (for overlap check)
    val_dates = plan_entry.get("validation_dates") or []
    if val_dates:
        begin = min(val_dates).replace("-", "")
    else:
        begin = plan_entry.get("fetch_start_date", "20240101").replace("-", "")
    end = datetime.now().strftime("%Y%m%d")

    freq_map = {"monthly": "月", "daily": "日", "quarterly": "季", "yearly": "年"}
    freq = freq_map.get(plan_entry.get("frequency", "monthly"), "月")

    # Try USD first, then CNY (matching how Excel/iFind stored the data)
    for currency in ("USD", "CNY"):
        result = wind_mcp_query(indicator, begin_date=begin, end_date=end,
                                freq=freq, magnitude="亿", currency=currency)
        if result and result.get("data"):
            return result["data"]
    return None


def simulate_fetch_from_excel(plan_entry, excel_path='FX Chartbook - Flow 0515.xlsx'):
    """
    Simulate a fetch by reading the latest data from the original Excel file.
    Used for pipeline testing when Wind MCP is unavailable.

    This extracts data AFTER the last DB date to simulate "new data arriving".
    """
    import openpyxl

    try:
        wb = openpyxl.load_workbook(excel_path, data_only=True)
    except FileNotFoundError:
        return None

    module = plan_entry.get('module', '')
    if module not in wb.sheetnames:
        return None

    ws = wb[module]
    col_letter = plan_entry['series_id'].split(':')[-1]
    try:
        col_idx = openpyxl.utils.column_index_from_string(col_letter)
    except ValueError:
        return None

    # Find data start row
    data_start = None
    for row_idx in range(1, 15):
        for c in range(1, 3):
            v = ws.cell(row=row_idx, column=c).value
            if v and '指标名称' in str(v):
                data_start = row_idx + 1  # Data starts after header+frequency rows
                break
        if data_start:
            break
    if not data_start:
        data_start = 6

    # Extract all data
    last_db_date = plan_entry.get('last_date', '')
    new_data = {}
    for row_idx in range(data_start, ws.max_row + 1):
        date_cell = ws.cell(row=row_idx, column=1)
        val_cell = ws.cell(row=row_idx, column=col_idx)

        if date_cell.value and val_cell.value is not None:
            if hasattr(date_cell.value, 'strftime'):
                date_str = date_cell.value.strftime('%Y-%m-%d')
            else:
                date_str = str(date_cell.value)[:10]

            # Only include dates after the last DB date (simulating new data)
            if date_str > last_db_date:
                try:
                    new_data[date_str] = float(val_cell.value)
                except (ValueError, TypeError):
                    pass

    return new_data if new_data else None


def main():
    import argparse
    parser = argparse.ArgumentParser(description="Fetch data from Wind MCP")
    parser.add_argument("--simulate", action="store_true",
                       help="Simulate fetch from Excel (for pipeline testing)")
    parser.add_argument("--plan", default="config/update_plan.json",
                       help="Path to update plan")
    parser.add_argument("--output", default="data/staging_fetched.json",
                       help="Path for staging output")
    args = parser.parse_args()

    plan_path = Path(args.plan)
    if not plan_path.exists():
        print(f"No update plan found at {plan_path}")
        print("Run build_update_plan.py first.")
        return 1

    with open(plan_path, 'r', encoding='utf-8') as f:
        plan = json.load(f)

    if not plan:
        print("Update plan is empty. Run build_update_plan.py with verified mappings.")
        return 1

    print(f"Fetching data for {len(plan)} series...")

    results = {}
    staging_records = {}
    fetched_count = 0
    failed_count = 0

    for entry in plan:
        sid = entry['series_id']

        if args.simulate:
            data = simulate_fetch_from_excel(entry)
        else:
            data = fetch_via_wind_mcp(entry)

        if data:
            # Apply controlled transform chain BEFORE staging (so overlap
            # validation downstream compares transformed values).
            chain = entry.get('transform') or []
            try:
                audit = apply_transforms_with_audit(data, chain)
            except ValueError as e:
                print(f"  ❌ {sid}: transform rejected — {e}")
                failed_count += 1
                continue

            transformed = audit['transformed_observations']
            results[sid] = transformed
            staging_records[sid] = {
                "series_id": sid,
                "query": entry.get('query', ''),
                "wind_code": entry.get('wind_code', ''),
                "wind_name": entry.get('wind_indicator', ''),
                "wind_unit": entry.get('unit', ''),
                "requested_frequency": entry.get('frequency', ''),
                "requested_currency": entry.get('currency', ''),
                "fetched_at": datetime.now().isoformat(),
                "raw_observations": audit['raw_observations'],
                "transform_chain": audit['transform_chain'],
                "transformed_observations": transformed,
            }
            fetched_count += 1
            print(f"  ✅ {sid}: {len(transformed)} new points, "
                  f"range {min(transformed.keys())} to {max(transformed.keys())}")
        else:
            failed_count += 1

    # Save staging file (full audit contract)
    staging = {
        "fetched_at": datetime.now().isoformat(),
        "method": "simulated" if args.simulate else "wind_mcp",
        "series_count": len(results),
        "series_data": results,
        "staging_records": staging_records,
    }

    output_path = Path(args.output)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with open(output_path, 'w', encoding='utf-8') as f:
        json.dump(staging, f, ensure_ascii=False, indent=2)

    print(f"\nFetched: {fetched_count} series, Failed: {failed_count}")
    print(f"Staging file: {output_path}")

    if not args.simulate and fetched_count == 0:
        print("\n⚠️  Wind MCP is not configured or returned no data.")
        print("To test the pipeline, run with --simulate to use Excel data.")

    return 0


if __name__ == "__main__":
    main()
