"""
Recompute derived indicators from raw observations.
Comprehensive recomputation engine for all 9 modules.

Architecture:
- Each derived series has formula_type, inputs, and optional params
- The engine aligns input series by date and executes the formula row-by-row
- Computed series are cached in-memory so dependents can use them in the same run
- After all definitions are processed, results are committed in a single batch
"""
import sys
from pathlib import Path
from datetime import datetime

sys.path.insert(0, str(Path(__file__).resolve().parent))
from lib import (
    get_db, init_db, start_update_run, finish_update_run,
    insert_observations_batch, upsert_series, get_observations
)


# =============================================================================
# Formula engines
# =============================================================================

def sma(values, window):
    """Simple moving average."""
    result = {}
    sorted_dates = sorted(values.keys())
    for i, d in enumerate(sorted_dates):
        if i >= window - 1:
            window_vals = [values[sorted_dates[j]] for j in range(i - window + 1, i + 1)]
            valid = [v for v in window_vals if v is not None]
            if len(valid) >= max(1, window // 2):
                result[d] = sum(valid) / len(valid)
    return result


def rolling_sum(values, window):
    """Rolling sum."""
    result = {}
    sorted_dates = sorted(values.keys())
    for i, d in enumerate(sorted_dates):
        if i >= window - 1:
            window_vals = [values[sorted_dates[j]] for j in range(i - window + 1, i + 1)]
            valid = [v for v in window_vals if v is not None]
            if len(valid) >= max(1, window // 2):
                result[d] = sum(valid)
    return result


def z_score(values, window=37, sample_std=True):
    """Rolling Z-score. sample_std=True matches Excel's STDEV (STDEV.S, N-1)."""
    result = {}
    sorted_dates = sorted(values.keys())
    for i, d in enumerate(sorted_dates):
        if i >= window - 1:
            window_vals = [values[sorted_dates[j]] for j in range(i - window + 1, i + 1)]
            valid = [v for v in window_vals if v is not None]
            if len(valid) >= max(1, window // 2):
                mean_v = sum(valid) / len(valid)
                n = len(valid)
                # Excel STDEV = STDEV.S = sample stddev (N-1)
                variance = sum((v - mean_v) ** 2 for v in valid) / (n - 1) if n > 1 and sample_std else sum((v - mean_v) ** 2 for v in valid) / n
                std_v = variance ** 0.5
                result[d] = (values[d] - mean_v) / std_v if std_v != 0 else 0.0
    return result


def rolling_sum_ratio(num_values, denom_values, window=3):
    """Ratio of rolling sums: SUM(num[0:window]) / SUM(denom[0:window])"""
    result = {}
    sorted_dates = sorted(set(num_values.keys()) & set(denom_values.keys()))
    for i, d in enumerate(sorted_dates):
        if i >= window - 1:
            num_vals = [num_values.get(sorted_dates[j]) for j in range(i - window + 1, i + 1)]
            den_vals = [denom_values.get(sorted_dates[j]) for j in range(i - window + 1, i + 1)]
            num_clean = [v for v in num_vals if v is not None]
            den_clean = [v for v in den_vals if v is not None]
            if len(num_clean) >= 2 and len(den_clean) >= 2:
                den_sum = sum(den_clean)
                if den_sum != 0:
                    result[d] = sum(num_clean) / den_sum
    return result


# =============================================================================
# Derived Indicator Definitions
# =============================================================================

DERIVED_DEFS = []

def _d(series_id, display_name, module, frequency, inputs, unit="", **params):
    DERIVED_DEFS.append({
        "series_id": series_id,
        "display_name": display_name,
        "module": module,
        "frequency": frequency,
        "unit": unit,
        "inputs": inputs,
        "params": params,
    })

# ---------------------------------------------------------------------------
# Module: 3.即远期
# ---------------------------------------------------------------------------

_d("fx_fwd:supply_demand", "外汇市场即远期总供求",
  "3.即远期", "monthly", ["fx_fwd:AB", "fx_fwd:AD", "fx_fwd:AJ"],
  unit="net_flow", fn="sum3")
_d("fx_fwd:spot_flow", "即期结售汇发生额（银行自身+代客）",
  "3.即远期", "monthly", ["fx_fwd:D", "fx_fwd:AB"],
  unit="net_flow", fn="sum2")
_d("fx_fwd:deriv_flow", "衍生品当月净签约（远期+期权）",
  "3.即远期", "monthly", ["fx_fwd:AD", "fx_fwd:AJ"],
  unit="net_flow", fn="sum2")
_d("fx_fwd:supply_demand_3mma", "外汇市场即远期总供求（3MMA）",
  "3.即远期", "monthly", ["fx_fwd:supply_demand"],
  unit="moving_average", fn="sma", window=3)
_d("fx_fwd:supply_demand_12mma", "外汇市场即远期总供求（12MMA）",
  "3.即远期", "monthly", ["fx_fwd:supply_demand"],
  unit="moving_average", fn="sma", window=12)
_d("fx_fwd:V", "即远期结售汇合计6MMA",
  "3.即远期", "monthly", ["fx_fwd:U"],
  unit="moving_average", fn="sma", window=6)
_d("fx_fwd:W", "即远期结售汇合计12MMA",
  "3.即远期", "monthly", ["fx_fwd:U"],
  unit="moving_average", fn="sma", window=12)
_d("fx_fwd:AP", "代客衍生品签约3MMA",
  "3.即远期", "monthly", ["fx_fwd:AO"],
  unit="moving_average", fn="sma", window=3)

# ---------------------------------------------------------------------------
# Module: 3.代客即期
# ---------------------------------------------------------------------------

# Level 1: Simple diffs (结汇 - 售汇 per category) — no dependencies
_d("fx_cspot:R", "货物贸易结售汇差额",
  "3.代客即期", "monthly", ["fx_cspot:C", "fx_cspot:J"],
  unit="net_flow", fn="diff")
_d("fx_cspot:S", "服务贸易结售汇差额",
  "3.代客即期", "monthly", ["fx_cspot:D", "fx_cspot:K"],
  unit="net_flow", fn="diff")
_d("fx_cspot:T", "收益和经常转移结售汇差额",
  "3.代客即期", "monthly", ["fx_cspot:E", "fx_cspot:L"],
  unit="net_flow", fn="diff")
_d("fx_cspot:U", "直接投资结售汇差额",
  "3.代客即期", "monthly", ["fx_cspot:G", "fx_cspot:N"],
  unit="net_flow", fn="diff")
_d("fx_cspot:V", "证券投资结售汇差额",
  "3.代客即期", "monthly", ["fx_cspot:H", "fx_cspot:O"],
  unit="net_flow", fn="diff")

# Level 2: W depends on U and V
# W = F - M - U - V = (资本与金融项目差额) - 直接投资差额 - 证券投资差额
_d("fx_cspot:W", "其他金融项目结售汇差额",
  "3.代客即期", "monthly",
  ["fx_cspot:F", "fx_cspot:M", "fx_cspot:U", "fx_cspot:V"],
  unit="net_flow", fn="custom_w")

# Level 3: X = R+S+T, Y = U+V+W, Z = X+Y — depend on R,S,T,U,V,W
_d("fx_cspot:X", "经常账户结售汇差额",
  "3.代客即期", "monthly",
  ["fx_cspot:R", "fx_cspot:S", "fx_cspot:T"],
  unit="net_flow", fn="sum3")
_d("fx_cspot:Y", "金融账户结售汇差额",
  "3.代客即期", "monthly",
  ["fx_cspot:U", "fx_cspot:V", "fx_cspot:W"],
  unit="net_flow", fn="sum3")
_d("fx_cspot:Z", "代客即期结售汇差额",
  "3.代客即期", "monthly",
  ["fx_cspot:X", "fx_cspot:Y"],
  unit="net_flow", fn="sum2")

# Level 4: Moving averages — depend on X, Y
_d("fx_cspot:AC", "经常账户结售汇6MMA",
  "3.代客即期", "monthly", ["fx_cspot:X"],
  unit="moving_average", fn="sma", window=6)
_d("fx_cspot:AD", "金融账户结售汇6MMA",
  "3.代客即期", "monthly", ["fx_cspot:Y"],
  unit="moving_average", fn="sma", window=6)
_d("fx_cspot:AE", "结售汇差额6MMA合计",
  "3.代客即期", "monthly", ["fx_cspot:AC", "fx_cspot:AD"],
  unit="moving_average", fn="sum2")

# ---------------------------------------------------------------------------
# Module: 3.涉外收付
# ---------------------------------------------------------------------------

# Level 1: Currency net flows
_d("fx_crossborder:W", "美元净流入",
  "3.涉外收付", "monthly", ["fx_crossborder:F", "fx_crossborder:H"],
  unit="net_flow", fn="diff")
_d("fx_crossborder:X", "人民币净流入",
  "3.涉外收付", "monthly", ["fx_crossborder:G", "fx_crossborder:I"],
  unit="net_flow", fn="diff")

# Level 2: Y = E - W - X (总顺差 - 美元净流入 - 人民币净流入)
_d("fx_crossborder:Y", "其他货币净流入",
  "3.涉外收付", "monthly",
  ["fx_crossborder:E", "fx_crossborder:W", "fx_crossborder:X"],
  unit="net_flow", fn="a_minus_b_minus_c")

# Level 3: Z = W + X + Y
_d("fx_crossborder:Z", "净流入合计",
  "3.涉外收付", "monthly",
  ["fx_crossborder:W", "fx_crossborder:X", "fx_crossborder:Y"],
  unit="net_flow", fn="sum3")

# Level 1 (raw-based): Category breakdowns
_d("fx_crossborder:AC", "货物贸易差额",
  "3.涉外收付", "monthly", ["fx_crossborder:L", "fx_crossborder:M"],
  unit="net_flow", fn="diff")
_d("fx_crossborder:AD", "服务贸易差额",
  "3.涉外收付", "monthly", ["fx_crossborder:N", "fx_crossborder:O"],
  unit="net_flow", fn="diff")
_d("fx_crossborder:AE", "其他经常项目差额",
  "3.涉外收付", "monthly",
  ["fx_crossborder:AI", "fx_crossborder:AC", "fx_crossborder:AD"],
  unit="net_flow", fn="a_minus_b_minus_c")
_d("fx_crossborder:AF", "直接投资差额",
  "3.涉外收付", "monthly", ["fx_crossborder:T", "fx_crossborder:U"],
  unit="net_flow", fn="diff")
_d("fx_crossborder:AG", "证券投资差额",
  "3.涉外收付", "monthly", ["fx_crossborder:R", "fx_crossborder:S"],
  unit="net_flow", fn="diff")
_d("fx_crossborder:AH", "其他金融项目差额",
  "3.涉外收付", "monthly",
  ["fx_crossborder:AJ", "fx_crossborder:AF", "fx_crossborder:AG"],
  unit="net_flow", fn="a_minus_b_minus_c")
_d("fx_crossborder:AI", "经常账户差额",
  "3.涉外收付", "monthly", ["fx_crossborder:J", "fx_crossborder:K"],
  unit="net_flow", fn="diff")
_d("fx_crossborder:AJ", "金融账户差额",
  "3.涉外收付", "monthly", ["fx_crossborder:P", "fx_crossborder:Q"],
  unit="net_flow", fn="diff")
_d("fx_crossborder:AK", "总顺差",
  "3.涉外收付", "monthly", ["fx_crossborder:B"],
  unit="net_flow", fn="copy")
_d("fx_crossborder:AL", "外币净流入(顺差-人民币净流入)",
  "3.涉外收付", "monthly",
  ["fx_crossborder:E", "fx_crossborder:X"],
  unit="net_flow", fn="diff")

# ---------------------------------------------------------------------------
# Module: 3.货物贸易
# ---------------------------------------------------------------------------

_d("trade_goods:K", "出口增速",
  "3.货物贸易", "monthly", ["trade_goods:B"],
  unit="yoy_pct", fn="export_growth")
_d("trade_goods:L", "出口增速3MMA",
  "3.货物贸易", "monthly", ["trade_goods:K"],
  unit="moving_average", fn="sma", window=3)
_d("trade_goods:R", "进出口差额",
  "3.货物贸易", "monthly", ["trade_goods:B", "trade_goods:E"],
  unit="net_flow", fn="diff")
_d("trade_goods:S", "涉外收付差额",
  "3.货物贸易", "monthly", ["trade_goods:C", "trade_goods:F"],
  unit="net_flow", fn="diff")
_d("trade_goods:T", "即期结汇差额",
  "3.货物贸易", "monthly", ["trade_goods:D", "trade_goods:G"],
  unit="net_flow", fn="diff")
_d("trade_goods:V", "顺差TTM",
  "3.货物贸易", "monthly", ["trade_goods:R"],
  unit="ttm", fn="sma", window=12)
_d("trade_goods:W", "流入TTM",
  "3.货物贸易", "monthly", ["trade_goods:S"],
  unit="ttm", fn="sma", window=12)
_d("trade_goods:X", "即期结汇TTM",
  "3.货物贸易", "monthly", ["trade_goods:T"],
  unit="ttm", fn="sma", window=12)
# Y = AVERAGE(J8:J19)*0.75
_d("trade_goods:Y", "假定75%为贸易商贡献",
  "3.货物贸易", "monthly", ["trade_goods:J"],
  unit="ttm", fn="sma_times_scalar", window=12, scalar=0.75)
_d("trade_goods:Z", "总结汇TTM",
  "3.货物贸易", "monthly", ["trade_goods:X", "trade_goods:Y"],
  unit="ttm", fn="sum2")
_d("trade_goods:AA", "滚动12个月 跨境-结汇(TTM)",
  "3.货物贸易", "monthly", ["trade_goods:W", "trade_goods:Z"],
  unit="ttm", fn="diff")
_d("trade_goods:AB", "滚动12个月 顺差-结汇(TTM)",
  "3.货物贸易", "monthly", ["trade_goods:V", "trade_goods:Z"],
  unit="ttm", fn="diff")

# ---------------------------------------------------------------------------
# Module: 3.贸易商
# ---------------------------------------------------------------------------

_d("trade_merchant:H", "顺差",
  "3.贸易商", "monthly", ["trade_merchant:B", "trade_merchant:E"],
  unit="net_flow", fn="diff")
_d("trade_merchant:K", "结汇/出口",
  "3.贸易商", "monthly", ["trade_merchant:D", "trade_merchant:B"],
  unit="ratio", fn="ratio")
_d("trade_merchant:L", "结汇/收入",
  "3.贸易商", "monthly", ["trade_merchant:D", "trade_merchant:C"],
  unit="ratio", fn="ratio")
_d("trade_merchant:M", "收入/出口",
  "3.贸易商", "monthly", ["trade_merchant:C", "trade_merchant:B"],
  unit="ratio", fn="ratio")
_d("trade_merchant:N", "购汇/进口",
  "3.贸易商", "monthly", ["trade_merchant:G", "trade_merchant:E"],
  unit="ratio", fn="ratio")
_d("trade_merchant:O", "购汇/支出",
  "3.贸易商", "monthly", ["trade_merchant:G", "trade_merchant:F"],
  unit="ratio", fn="ratio")
_d("trade_merchant:P", "支出/进口",
  "3.贸易商", "monthly", ["trade_merchant:F", "trade_merchant:E"],
  unit="ratio", fn="ratio")

# Z-Score (37-period rolling)
_d("trade_merchant:Q", "货物贸易收汇结汇率 Z-Score",
  "3.贸易商", "monthly", ["trade_merchant:L"],
  unit="z_score", fn="zscore", window=37)
_d("trade_merchant:R", "货物贸易付汇购汇率 Z-Score",
  "3.贸易商", "monthly", ["trade_merchant:O"],
  unit="z_score", fn="zscore", window=37)

# 3M rolling sum ratios
_d("trade_merchant:T", "贸易顺差净结汇意愿 3M",
  "3.贸易商", "monthly", ["trade_merchant:D", "trade_merchant:B"],
  unit="ratio_3m", fn="rolling_sum_ratio", window=3)
_d("trade_merchant:U", "收付顺差结汇意愿 3M",
  "3.贸易商", "monthly", ["trade_merchant:D", "trade_merchant:C"],
  unit="ratio_3m", fn="rolling_sum_ratio", window=3)
_d("trade_merchant:V", "购汇/进口 3M",
  "3.贸易商", "monthly", ["trade_merchant:G", "trade_merchant:E"],
  unit="ratio_3m", fn="rolling_sum_ratio", window=3)
_d("trade_merchant:W", "购汇/支出 3M",
  "3.贸易商", "monthly", ["trade_merchant:G", "trade_merchant:F"],
  unit="ratio_3m", fn="rolling_sum_ratio", window=3)

# Level 2: Depend on above
_d("trade_merchant:X", "收付顺差净结汇意愿",
  "3.贸易商", "monthly", ["trade_merchant:U", "trade_merchant:W"],
  unit="ratio", fn="diff")
_d("trade_merchant:Y", "贸易顺差净结汇意愿",
  "3.贸易商", "monthly", ["trade_merchant:T", "trade_merchant:W"],
  unit="ratio", fn="diff")

_d("trade_merchant:AL", "进出口差额",
  "3.贸易商", "monthly", ["trade_merchant:B", "trade_merchant:E"],
  unit="net_flow", fn="diff")
_d("trade_merchant:AM", "涉外收付差额",
  "3.贸易商", "monthly", ["trade_merchant:C", "trade_merchant:F"],
  unit="net_flow", fn="diff")
_d("trade_merchant:AN", "结售汇差额",
  "3.贸易商", "monthly", ["trade_merchant:D", "trade_merchant:G"],
  unit="net_flow", fn="diff")
_d("trade_merchant:AU", "结售汇差额3MMA",
  "3.贸易商", "monthly", ["trade_merchant:AN"],
  unit="moving_average", fn="sma", window=3)
_d("trade_merchant:AV", "进出口差额3MMA",
  "3.贸易商", "monthly", ["trade_merchant:AL"],
  unit="moving_average", fn="sma", window=3)
_d("trade_merchant:AW", "涉外收付差额3MMA",
  "3.贸易商", "monthly", ["trade_merchant:AM"],
  unit="moving_average", fn="sma", window=3)
_d("trade_merchant:AX", "滚动3个月净结汇比例",
  "3.贸易商", "monthly", ["trade_merchant:AU", "trade_merchant:AV"],
  unit="ratio", fn="ratio")
_d("trade_merchant:AY", "当月净结汇比例",
  "3.贸易商", "monthly", ["trade_merchant:AN", "trade_merchant:AL"],
  unit="ratio", fn="ratio")

# ---------------------------------------------------------------------------
# Module: 3.FDI
# ---------------------------------------------------------------------------

_d("fdi:Q", "对外直接投资",
  "3.FDI", "monthly", ["fdi:G"], unit="value", fn="copy")
_d("fdi:T", "FDI结汇3MMA",
  "3.FDI", "monthly", ["fdi:I"], unit="moving_average", fn="sma", window=3)
# U = -AVERAGE(J6:J8) — negated 3-month SMA of column J (售汇)
_d("fdi:U", "FDI购汇3MMA",
  "3.FDI", "monthly", ["fdi:J"], unit="moving_average",
  fn="sma_times_scalar", window=3, scalar=-1)
# S = T + U where T=sma(I,3), U=-sma(J,3) → S = sma(I,3) - sma(J,3)
_d("fdi:S", "FDI结售汇差额3MMA",
  "3.FDI", "monthly", ["fdi:T", "fdi:U"],
  unit="moving_average", fn="sum2")

_d("fdi:P", "实际使用外资",
  "3.FDI", "monthly", ["fdi:B", "fdi:K", "fdi:M"],
  unit="value", fn="fdi_P")
_d("fdi:R", "FDI-ODI差额",
  "3.FDI", "monthly", ["fdi:P", "fdi:Q"],
  unit="net_flow", fn="diff")
_d("fdi:AY", "FDI流入-股权",
  "3.FDI", "monthly", ["fdi:AR"], unit="value", fn="copy")
_d("fdi:AZ", "FDI流入-债权",
  "3.FDI", "monthly", ["fdi:AS"], unit="value", fn="copy")

# ---------------------------------------------------------------------------
# Module: 3.证券FI
# ---------------------------------------------------------------------------

_d("sec_fi:AH", "国债持仓",
  "3.证券FI", "monthly", ["sec_fi:B"], unit="stock", fn="copy")
_d("sec_fi:AJ", "同业存单持仓",
  "3.证券FI", "monthly", ["sec_fi:P"], unit="stock", fn="copy")
# AI = IF((C+D+E)=0,W,C+D+E) — 政金债
_d("sec_fi:AI", "政金债持仓",
  "3.证券FI", "monthly", ["sec_fi:C", "sec_fi:D", "sec_fi:E", "sec_fi:W"],
  unit="stock", fn="secfi_AI")

# MoM changes (stock → flow) — depend on above
_d("sec_fi:AS", "国债持仓变动",
  "3.证券FI", "monthly", ["sec_fi:AH"], unit="flow", fn="mom_diff_skip_zero")
_d("sec_fi:AT", "政金债持仓变动",
  "3.证券FI", "monthly", ["sec_fi:AI"], unit="flow", fn="mom_diff_skip_zero")
_d("sec_fi:AU", "同业存单持仓变动",
  "3.证券FI", "monthly", ["sec_fi:AJ"], unit="flow", fn="mom_diff_skip_zero")

_d("sec_fi:BA", "利率债合计变动",
  "3.证券FI", "monthly", ["sec_fi:AS", "sec_fi:AT"],
  unit="flow", fn="sum2")
_d("sec_fi:BB", "国债+政金债变动3MMA",
  "3.证券FI", "monthly", ["sec_fi:BA"],
  unit="moving_average", fn="sma", window=3)
_d("sec_fi:BK", "利率债inflow",
  "3.证券FI", "monthly", ["sec_fi:BB"], unit="flow", fn="copy")
_d("sec_fi:BJ", "中美利差",
  "3.证券FI", "monthly", ["sec_fi:BH", "sec_fi:BI"],
  unit="bp", fn="diff")

# ---------------------------------------------------------------------------
# Module: 3.证券EQ (daily)
# ---------------------------------------------------------------------------

_d("sec_eq:P", "陆港通净流入",
  "3.证券EQ", "daily", ["sec_eq:O", "sec_eq:N"],
  unit="net_flow", fn="diff")


# =============================================================================
# Formula function implementations
# =============================================================================

def execute_fn(fn_name, input_data, params, all_dates, series_cache):
    """Execute a named formula function against aligned input data.

    input_data: dict of {idx: {date: value}} where idx is 0, 1, 2, ...
    Returns {date: value}
    """
    result = {}

    def get(idx, default=None):
        return input_data.get(idx, default or {})

    if fn_name == "copy":
        d0 = get(0)
        for d in all_dates:
            if d in d0 and d0[d] is not None:
                result[d] = d0[d]

    elif fn_name == "negate":
        d0 = get(0)
        for d in all_dates:
            if d in d0 and d0[d] is not None:
                result[d] = -d0[d]

    elif fn_name == "diff":
        d0, d1 = get(0), get(1)
        for d in all_dates:
            v0, v1 = d0.get(d), d1.get(d)
            if v0 is not None and v1 is not None:
                result[d] = v0 - v1

    elif fn_name == "sum2":
        d0, d1 = get(0), get(1)
        for d in all_dates:
            v0, v1 = d0.get(d), d1.get(d)
            if v0 is not None and v1 is not None:
                result[d] = v0 + v1

    elif fn_name == "sum3":
        d0, d1, d2 = get(0), get(1), get(2)
        for d in all_dates:
            v0, v1, v2 = d0.get(d), d1.get(d), d2.get(d)
            if v0 is not None and v1 is not None and v2 is not None:
                result[d] = v0 + v1 + v2

    elif fn_name == "ratio":
        d0, d1 = get(0), get(1)
        for d in all_dates:
            v0, v1 = d0.get(d), d1.get(d)
            if v0 is not None and v1 is not None and v1 != 0:
                result[d] = v0 / v1

    elif fn_name == "sma":
        d0 = get(0)
        window = params.get("window", 3)
        result = sma(d0, window)

    elif fn_name == "sma_times_scalar":
        d0 = get(0)
        window = params.get("window", 3)
        scalar = params.get("scalar", 1.0)
        temp = sma(d0, window)
        for d, v in temp.items():
            result[d] = v * scalar

    elif fn_name == "zscore":
        d0 = get(0)
        window = params.get("window", 37)
        result = z_score(d0, window)

    elif fn_name == "rolling_sum_ratio":
        d0, d1 = get(0), get(1)
        window = params.get("window", 3)
        result = rolling_sum_ratio(d0, d1, window)

    elif fn_name == "a_minus_b_minus_c":
        d0, d1, d2 = get(0), get(1), get(2)
        for d in all_dates:
            v0, v1, v2 = d0.get(d), d1.get(d), d2.get(d)
            if v0 is not None and v1 is not None and v2 is not None:
                result[d] = v0 - v1 - v2

    elif fn_name == "custom_w":
        # W = F - M - U - V (资本与金融项目 - 直接投资 - 证券投资)
        d0, d1, d2, d3 = get(0), get(1), get(2), get(3)
        for d in all_dates:
            v0, v1, v2, v3 = d0.get(d), d1.get(d), d2.get(d), d3.get(d)
            if all(v is not None for v in [v0, v1, v2, v3]):
                result[d] = v0 - v1 - v2 - v3

    elif fn_name == "export_growth":
        d0 = get(0)
        for d in all_dates:
            va = d0.get(d)
            if va is not None and va != 0:
                year = int(d[:4])
                month = int(d[5:7])
                day = d[8:10]  # Extract day from current date
                prev_date = f"{year-1}-{month:02d}-{day}"
                vb = d0.get(prev_date)
                if year == 2021:
                    prev2_date = f"{year-2}-{month:02d}-{day}"
                    vc = d0.get(prev2_date)
                    if vb is not None and vc is not None and vc != 0:
                        result[d] = (va / vc - 1) / 2
                elif vb is not None and vb != 0:
                    result[d] = va / vb - 1

    elif fn_name == "mom_diff_skip_zero":
        # MoM change: v[n] - v[n-1], skip if v[n]==0
        d0 = get(0)
        sorted_dates = sorted(d0.keys())
        for i, d in enumerate(sorted_dates):
            if i > 0:
                curr = d0[d]
                prev = d0[sorted_dates[i - 1]]
                if curr is not None and curr != 0 and prev is not None:
                    result[d] = curr - prev
                elif curr == 0:
                    result[d] = 0.0

    elif fn_name == "fdi_P":
        # IF(B=0, IF(K=0, #N/A, K/M), B)
        d0, d1, d2 = get(0), get(1), get(2)
        for d in all_dates:
            b = d0.get(d)
            k = d1.get(d)
            m = d2.get(d)
            if b is not None and b != 0:
                result[d] = b
            elif b == 0 and k is not None and m is not None and m != 0:
                result[d] = k / m

    elif fn_name == "secfi_AI":
        # IF((C+D+E)=0, W, C+D+E)
        d0, d1, d2, d3 = get(0), get(1), get(2), get(3)
        for d in all_dates:
            c, de, e, w = d0.get(d), d1.get(d), d2.get(d), d3.get(d)
            if c is not None and de is not None and e is not None:
                s = c + de + e
                if s == 0 and w is not None:
                    result[d] = w
                elif s != 0:
                    result[d] = s

    return result


# =============================================================================
# Main recomputation logic
# =============================================================================

def recompute_all(conn, run_id, module_filter=None):
    """Recompute all derived indicators. Returns (computed_ids, skipped_ids, n_obs)."""
    series_cache = {}

    def load_series(sid):
        if sid not in series_cache:
            rows = get_observations(conn, sid)
            series_cache[sid] = {r["date"]: r["value"] for r in rows}
        return series_cache[sid]

    # Pre-load all input series
    all_input_sids = set()
    for defn in DERIVED_DEFS:
        if module_filter and defn["module"] != module_filter:
            continue
        for inp in defn["inputs"]:
            all_input_sids.add(inp)

    print(f"Pre-loading {len(all_input_sids)} input series...")
    for sid in sorted(all_input_sids):
        try:
            load_series(sid)
        except Exception as e:
            print(f"  WARNING: Cannot load {sid}: {e}")

    all_new_obs = []
    computed = []
    skipped = []

    for defn in DERIVED_DEFS:
        if module_filter and defn["module"] != module_filter:
            continue

        sid = defn["series_id"]
        fn_name = defn.get("params", {}).get("fn", "sum2")
        display = defn["display_name"]
        print(f"  Computing: {sid} — {display} ({fn_name})")

        # Load input data from cache
        input_data = {}
        all_dates = set()
        for idx, inp_sid in enumerate(defn["inputs"]):
            if inp_sid in series_cache:
                data = series_cache[inp_sid]
            else:
                data = load_series(inp_sid)
            input_data[idx] = data
            all_dates.update(data.keys())

        all_dates = sorted(all_dates)

        if not all_dates:
            print(f"    WARNING: No dates found for inputs {defn['inputs']}")
            skipped.append(sid)
            continue

        # Execute formula
        result = execute_fn(
            fn_name, input_data, defn.get("params", {}), all_dates, series_cache
        )

        if not result:
            print(f"    WARNING: No values computed")
            skipped.append(sid)
            continue

        # Store result in cache for downstream dependents
        series_cache[sid] = result

        sorted_result = sorted(result.keys())
        first_date = sorted_result[0]
        last_date = sorted_result[-1]

        formula_desc = f"{fn_name}: inputs={defn['inputs']} params={defn.get('params', {})}"

        upsert_series(conn, {
            "series_id": sid,
            "display_name": display,
            "module": defn["module"],
            "series_type": "derived",
            "frequency": defn["frequency"],
            "unit": defn.get("unit", ""),
            "source": "recomputed",
            "source_query": formula_desc,
            "excel_sheet": defn["module"],
            "update_status": "computed",
            "first_date": first_date,
            "last_date": last_date,
            "notes": f"Inputs: {', '.join(defn['inputs'])}. fn={fn_name}",
        })

        for d, v in result.items():
            all_new_obs.append((sid, d, v, "recomputed"))

        computed.append(sid)
        print(f"    {len(result)} values, range {first_date} to {last_date}")

    # Batch insert
    if all_new_obs:
        insert_observations_batch(conn, all_new_obs, run_id)
        conn.commit()
        print(f"  Inserted {len(all_new_obs)} derived observations")

    if skipped:
        print(f"  Skipped {len(skipped)}: {skipped}")

    return computed, skipped, len(all_new_obs)


def compare_with_excel(conn, computed_ids, sheet_name, data_start):
    """Compare recomputed values against Excel file."""
    import openpyxl
    wb = openpyxl.load_workbook('FX Chartbook - Flow 0515.xlsx', data_only=True)

    if sheet_name not in wb.sheetnames:
        print(f"  Sheet '{sheet_name}' not found in workbook")
        return {}

    ws = wb[sheet_name]
    results = {}

    for sid in computed_ids:
        col_letter = sid.split(':')[-1]
        try:
            col_idx = openpyxl.utils.column_index_from_string(col_letter)
        except ValueError:
            continue

        recomputed = dict(get_observations(conn, sid))

        # Build Excel map
        excel_map = {}
        for row_idx in range(data_start, ws.max_row + 1):
            date_cell = ws.cell(row=row_idx, column=1)
            val_cell = ws.cell(row=row_idx, column=col_idx)
            if date_cell.value and val_cell.value is not None:
                if hasattr(date_cell.value, 'strftime'):
                    date_str = date_cell.value.strftime('%Y-%m-%d')
                else:
                    date_str = str(date_cell.value)[:10]
                try:
                    excel_map[date_str] = float(val_cell.value)
                except (ValueError, TypeError):
                    pass

        matches, mismatches, max_diff = 0, 0, 0.0
        max_diff_info = ""
        for d, rv in recomputed.items():
            ev = excel_map.get(d)
            if ev is not None:
                diff = abs(rv - ev)
                if diff < 1e-6 or (abs(ev) > 1e-9 and diff / abs(ev) < 1e-6):
                    matches += 1
                else:
                    mismatches += 1
                    if diff > max_diff:
                        max_diff = diff
                        max_diff_info = f"{sid} {d}: py={rv:.6f} excel={ev:.6f} diff={diff:.6f}"

        results[sid] = {
            "matches": matches, "mismatches": mismatches,
            "max_diff": max_diff, "max_diff_info": max_diff_info,
            "total": matches + mismatches,
        }

    return results


def main():
    import argparse
    parser = argparse.ArgumentParser(description="Recompute derived indicators")
    parser.add_argument("--module", default=None, help="Module to recompute (omit for all)")
    parser.add_argument("--compare", action="store_true", help="Compare with Excel after recompute")
    args = parser.parse_args()

    MODULE_SHEET_MAP = {
        "3.即远期": ("3.即远期", 7),
        "3.代客即期": ("3.代客即期", 6),
        "3.涉外收付": ("3.涉外收付", 6),
        "3.货物贸易": ("3.货物贸易", 8),
        "3.贸易商": ("3.贸易商", 4),
        "3.FDI": ("3.FDI", 6),
        "3.证券FI": ("3.证券FI", 4),
        "3.证券EQ": ("3.证券EQ", 12),
    }

    conn = get_db()
    init_db(conn)
    run_id = start_update_run(conn, 0)

    module_str = args.module or "all modules"
    print(f"Recomputing derived indicators for: {module_str}")
    print(f"Total definitions: {len(DERIVED_DEFS)}")

    computed, skipped, n_obs = recompute_all(conn, run_id, args.module)

    finish_update_run(conn, run_id, "completed",
                      successful=len(computed),
                      new_obs=n_obs)

    print(f"\nDone: {len(computed)} computed, {len(skipped)} skipped, {n_obs} obs")

    if args.compare and computed:
        by_module = {}
        for sid in computed:
            for defn in DERIVED_DEFS:
                if defn["series_id"] == sid:
                    mod = defn["module"]
                    by_module.setdefault(mod, []).append(sid)
                    break

        total_match, total_mismatch = 0, 0
        for mod, sids in by_module.items():
            sheet_info = MODULE_SHEET_MAP.get(mod)
            if not sheet_info:
                continue
            print(f"\n--- {mod} vs Excel ---")
            results = compare_with_excel(conn, sids, *sheet_info)
            for sid, r in results.items():
                status = "✅" if r["mismatches"] == 0 else "❌"
                print(f"  {status} {sid}: {r['matches']} match, {r['mismatches']} mismatch, "
                      f"max_diff={r['max_diff']:.6f}")
                total_match += r["matches"]
                total_mismatch += r["mismatches"]

        pct = total_match / (total_match + total_mismatch) * 100 if (total_match + total_mismatch) > 0 else 0
        print(f"\nTotal: {total_match} match, {total_mismatch} mismatch ({pct:.1f}%)")

    conn.close()
    print("Done.")


if __name__ == "__main__":
    main()
