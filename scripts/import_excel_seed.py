"""
Import Excel seed data into SQLite.
Reads cached values from Excel, imports as historical baseline.
Idempotent: re-running produces no duplicate observations.
"""
import sys
import argparse
from pathlib import Path
from datetime import datetime, date
import openpyxl
from openpyxl.utils import get_column_letter

# Add project root to path
sys.path.insert(0, str(Path(__file__).resolve().parent))
from lib import (
    get_db, init_db, upsert_series, insert_observations_batch,
    start_update_run, finish_update_run, DB_PATH
)

WORKBOOK = Path("/Users/martin_ai/Desktop/Martin Monthly Brief/FX Chartbook - Flow 0515.xlsx")

# Module prefix mapping
MODULE_PREFIX = {
    "3.即远期": "fx_fwd",
    "3.代客即期": "fx_cspot",
    "3.涉外收付": "fx_crossborder",
    "3.货物贸易": "trade_goods",
    "3.贸易商": "trade_merchant",
    "3.服务贸易": "trade_services",
    "3.FDI": "fdi",
    "3.证券EQ": "sec_eq",
    "3.证券FI": "sec_fi",
}

# Frequency mapping
MODULE_FREQ = {
    "3.即远期": "monthly",
    "3.代客即期": "monthly",
    "3.涉外收付": "monthly",
    "3.货物贸易": "monthly",
    "3.贸易商": "monthly",
    "3.服务贸易": "monthly",
    "3.FDI": "monthly",
    "3.证券EQ": "daily",
    "3.证券FI": "monthly",
}


def safe_val(cell):
    """Extract cell value safely."""
    if cell is None or cell.value is None:
        return None
    if isinstance(cell.value, str):
        return cell.value.strip()
    return cell.value


def is_formula(cell):
    """Check if cell contains a formula."""
    if cell is None:
        return False
    return isinstance(cell.value, str) and str(cell.value).startswith("=")


def parse_date(val):
    """Parse a date value to ISO string."""
    if isinstance(val, datetime):
        return val.strftime("%Y-%m-%d")
    if isinstance(val, date):
        return val.isoformat()
    return None


def determine_unit(header_text, col_letter):
    """Infer unit from column header text."""
    if not header_text:
        return "unknown"
    text = str(header_text)
    if "当月值" in text or "当月" in text:
        return "monthly_amount"
    if "累计值" in text or "累计" in text:
        return "cumulative_amount"
    if "累计同比" in text or "同比" in text:
        return "yoy_pct"
    if "净流入" in text or "净结汇" in text:
        return "net_flow"
    if "占比" in text or "比重" in text:
        return "percentage"
    if "指数" in text:
        return "index"
    if "收益率" in text:
        return "yield_pct"
    if "差额" in text or "顺差" in text:
        return "balance"
    if "mma" in text.lower():
        return "moving_average"
    if "z-score" in text.lower():
        return "z_score"
    if "correl" in text.lower():
        return "correlation"
    if "增速" in text or "change" in text.lower():
        return "pct_change"
    if "汇率" in text:
        return "exchange_rate"
    if "托管量" in text:
        return "custody_amount"
    if "净买入" in text:
        return "net_buy"
    return "unknown"


def import_sheet(conn, sheet_name, run_id):
    """Import all data from a single sheet."""
    prefix = MODULE_PREFIX.get(sheet_name, sheet_name.replace("3.", ""))
    frequency = MODULE_FREQ.get(sheet_name, "monthly")

    print(f"  Opening workbook for {sheet_name}...")
    # Read values (cached) and formulas separately
    wb_v = openpyxl.load_workbook(WORKBOOK, data_only=True, read_only=False)
    wb_f = openpyxl.load_workbook(WORKBOOK, data_only=False, read_only=False)

    ws_v = wb_v[sheet_name]
    ws_f = wb_f[sheet_name]

    # Step 1: Find the date column and data start row
    date_col_idx = None
    data_start_row = None

    for row_idx in range(1, min(31, ws_v.max_row + 1)):
        for col_idx in range(1, min(10, ws_v.max_column + 1)):
            cell = ws_v.cell(row=row_idx, column=col_idx)
            val = safe_val(cell)
            if isinstance(val, (datetime, date)):
                date_col_idx = col_idx
                data_start_row = row_idx
                break
        if date_col_idx:
            break

    if date_col_idx is None:
        print(f"  WARNING: Could not find date column for {sheet_name}, trying column A")
        date_col_idx = 1
        data_start_row = 6  # Common pattern

    print(f"  Date column: {get_column_letter(date_col_idx)}, data starts at row {data_start_row}")

    # Step 2: Extract dates
    dates = []
    for row_idx in range(data_start_row, ws_v.max_row + 1):
        cell = ws_v.cell(row=row_idx, column=date_col_idx)
        val = safe_val(cell)
        if isinstance(val, (datetime, date)):
            dates.append((row_idx, val))
        elif val is not None and row_idx <= data_start_row + 10:
            continue  # Skip non-date in first few rows

    if not dates:
        print("  ERROR: No dates found!")
        wb_v.close()
        wb_f.close()
        return 0

    print(f"  Found {len(dates)} dates: {parse_date(dates[0][1])} to {parse_date(dates[-1][1])}")

    # Step 3: Find the header row (row with most text in first ~30 rows)
    header_row = None
    max_text = 0
    for row_idx in range(1, min(31, ws_v.max_row + 1)):
        text_count = 0
        for col_idx in range(1, min(ws_v.max_column + 1, 70)):
            cell = ws_v.cell(row=row_idx, column=col_idx)
            val = safe_val(cell)
            if isinstance(val, str) and len(val) > 1:
                text_count += 1
        if text_count > max_text and text_count >= 2:
            max_text = text_count
            header_row = row_idx

    if header_row is None:
        header_row = data_start_row

    print(f"  Header row: {header_row}")

    # Step 4: Import each column as a series
    imported_series = 0
    total_observations = 0
    all_obs = []

    for col_idx in range(1, min(ws_v.max_column + 1, 70)):
        col_letter = get_column_letter(col_idx)
        series_id = f"{prefix}:{col_letter}"

        # Get header
        header_val = safe_val(ws_v.cell(row=header_row, column=col_idx))
        if header_val is None or (isinstance(header_val, str) and len(header_val) == 0):
            # Try row 1 or data_start_row-1 as fallback
            for fallback_row in [1, data_start_row - 1, data_start_row - 2]:
                if fallback_row > 0:
                    fallback = safe_val(ws_v.cell(row=fallback_row, column=col_idx))
                    if fallback and isinstance(fallback, str) and len(fallback) > 1:
                        header_val = fallback
                        break

        # Determine if this column has formulas
        has_formulas = False
        for row_idx, _ in dates[:10]:
            if is_formula(ws_f.cell(row=row_idx, column=col_idx)):
                has_formulas = True
                break

        # Determine series_type
        if col_idx == date_col_idx:
            series_type = "raw"  # Date column
        elif has_formulas:
            series_type = "derived"
        else:
            # Check if it has numeric data
            numeric_count = 0
            for row_idx, _ in dates[:20]:
                val = safe_val(ws_v.cell(row=row_idx, column=col_idx))
                if isinstance(val, (int, float)):
                    numeric_count += 1
            series_type = "raw" if numeric_count > 0 else "manual"

        # Extract observations
        col_obs = []
        first_date = None
        last_date = None

        for row_idx, dt in dates:
            cell = ws_v.cell(row=row_idx, column=col_idx)
            val = safe_val(cell)
            if isinstance(val, (int, float)):
                date_str = parse_date(dt)
                if date_str:
                    col_obs.append((series_id, date_str, float(val), "excel_seed"))
                    if first_date is None:
                        first_date = date_str
                    last_date = date_str

        # Skip columns with too few observations
        if len(col_obs) < 5 and col_idx != date_col_idx:
            continue

        # Upsert series metadata
        unit = determine_unit(str(header_val) if header_val else "", col_letter)
        upsert_series(conn, {
            "series_id": series_id,
            "display_name": str(header_val)[:200] if header_val else f"Column_{col_letter}",
            "module": sheet_name,
            "series_type": series_type,
            "frequency": frequency,
            "unit": unit,
            "source": "excel_seed",
            "excel_sheet": sheet_name,
            "excel_range": f"{col_letter}{data_start_row}:{col_letter}{ws_v.max_row}",
            "update_status": "imported",
            "first_date": first_date,
            "last_date": last_date,
            "notes": f"Imported from Excel, {len(col_obs)} observations",
        })

        all_obs.extend(col_obs)
        imported_series += 1
        total_observations += len(col_obs)

    # Step 5: Batch insert observations
    if all_obs:
        insert_observations_batch(conn, all_obs, run_id)
        conn.commit()

    wb_v.close()
    wb_f.close()

    print(f"  Imported {imported_series} series, {total_observations} observations")
    return imported_series


def main():
    parser = argparse.ArgumentParser(description="Import Excel seed data to SQLite")
    parser.add_argument("--module", help="Sheet name to import (default: all 9 target sheets)")
    parser.add_argument("--all", action="store_true", help="Import all 9 target sheets")
    args = parser.parse_args()

    # Determine which sheets to import
    if args.module:
        sheets = [args.module]
    elif args.all:
        sheets = list(MODULE_PREFIX.keys())
    else:
        # Default: import all 9 sheets
        sheets = list(MODULE_PREFIX.keys())

    print(f"Importing from: {WORKBOOK}")
    print(f"Target sheets: {sheets}")
    print(f"Output DB: {DB_PATH}")

    # Initialize
    conn = get_db()
    init_db(conn)
    run_id = start_update_run(conn, len(sheets))

    total = 0
    for sheet_name in sheets:
        print(f"\n--- {sheet_name} ---")
        try:
            n = import_sheet(conn, sheet_name, run_id)
            total += n
        except Exception as e:
            print(f"  ERROR importing {sheet_name}: {e}")
            import traceback
            traceback.print_exc()

    finish_update_run(conn, run_id, "completed" if total > 0 else "failed",
                      successful=total, new_obs=0)

    # Quick validation
    series_count = conn.execute("SELECT COUNT(*) as cnt FROM series").fetchone()["cnt"]
    obs_count = conn.execute("SELECT COUNT(*) as cnt FROM observations").fetchone()["cnt"]
    conn.close()

    print(f"\n{'='*60}")
    print(f"Import complete: {total} series, {series_count} total series in DB, {obs_count} total observations")


if __name__ == "__main__":
    main()
