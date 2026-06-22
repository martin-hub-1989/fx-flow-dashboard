import json
import re
import sys
from collections import Counter
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


def safe_value(value, limit=180):
    if value is None:
        return None
    if hasattr(value, "isoformat"):
        try:
            return value.isoformat()
        except Exception:
            pass
    text = str(value)
    return text if len(text) <= limit else text[: limit - 3] + "..."


def chart_title(chart):
    try:
        title = chart.title
        if title is None:
            return None
        parts = []
        for paragraph in title.tx.rich.p:
            for run in paragraph.r:
                if run.t:
                    parts.append(run.t)
            if paragraph.endParaRPr and getattr(paragraph.endParaRPr, "t", None):
                parts.append(paragraph.endParaRPr.t)
        return "".join(parts) or str(title)
    except Exception:
        return None


def cell_matrix(ws, min_row, max_row, min_col, max_col):
    rows = []
    for row in ws.iter_rows(
        min_row=min_row,
        max_row=max_row,
        min_col=min_col,
        max_col=max_col,
    ):
        rows.append([safe_value(cell.value) for cell in row])
    return rows


def main():
    workbook_path = Path(sys.argv[1])
    output_path = Path(sys.argv[2])
    wb = load_workbook(workbook_path, data_only=False, read_only=False)

    report = {
        "workbook": str(workbook_path),
        "sheetnames": wb.sheetnames,
        "defined_names": [str(item) for item in wb.defined_names.values()],
        "external_links": len(getattr(wb, "_external_links", [])),
        "sheets": [],
    }

    formula_ref_counter = Counter()
    formula_function_counter = Counter()

    for ws in wb.worksheets:
        formulas = []
        wind_like = []
        labels = []
        nonempty = 0
        first_nonempty_row = None
        last_nonempty_row = 0
        first_nonempty_col = None
        last_nonempty_col = 0

        for row in ws.iter_rows():
            for cell in row:
                value = cell.value
                if value is None:
                    continue
                nonempty += 1
                first_nonempty_row = cell.row if first_nonempty_row is None else min(first_nonempty_row, cell.row)
                last_nonempty_row = max(last_nonempty_row, cell.row)
                first_nonempty_col = cell.column if first_nonempty_col is None else min(first_nonempty_col, cell.column)
                last_nonempty_col = max(last_nonempty_col, cell.column)
                text = str(value)
                if cell.data_type == "f" or text.startswith("="):
                    formulas.append({"cell": cell.coordinate, "formula": safe_value(text, 500)})
                    for fn in re.findall(r"\b([A-Z][A-Z0-9._]+)\s*\(", text.upper()):
                        formula_function_counter[fn] += 1
                    for ref in re.findall(r"'([^']+)'!", text):
                        formula_ref_counter[ref] += 1
                    if any(token in text.upper() for token in ("WIND", "WSD", "WSS", "EDB", "WSET", "WPF")):
                        wind_like.append({"cell": cell.coordinate, "value": safe_value(text, 500)})
                elif isinstance(value, str):
                    if any(token in value.upper() for token in ("WIND", "WSD", "WSS", "EDB", "WSET", "WPF")):
                        wind_like.append({"cell": cell.coordinate, "value": safe_value(value, 500)})
                    if len(labels) < 250 and value.strip():
                        labels.append({"cell": cell.coordinate, "value": safe_value(value)})

        min_row = first_nonempty_row or 1
        max_row = min(last_nonempty_row or 1, min_row + 14)
        min_col = first_nonempty_col or 1
        max_col = min(last_nonempty_col or 1, min_col + 19)

        charts = []
        for idx, chart in enumerate(getattr(ws, "_charts", []), start=1):
            anchor = getattr(chart, "anchor", None)
            anchor_desc = None
            try:
                anchor_desc = {
                    "from": {
                        "row": anchor._from.row + 1,
                        "col": anchor._from.col + 1,
                    },
                    "to": {
                        "row": anchor.to.row + 1,
                        "col": anchor.to.col + 1,
                    },
                }
            except Exception:
                anchor_desc = safe_value(anchor)
            charts.append({
                "index": idx,
                "type": type(chart).__name__,
                "title": chart_title(chart),
                "style": getattr(chart, "style", None),
                "anchor": anchor_desc,
                "series_count": len(getattr(chart, "ser", [])),
            })

        report["sheets"].append({
            "name": ws.title,
            "state": ws.sheet_state,
            "dimensions": ws.calculate_dimension(),
            "max_row": ws.max_row,
            "max_column": ws.max_column,
            "nonempty_cells": nonempty,
            "nonempty_bounds": {
                "min_row": first_nonempty_row,
                "max_row": last_nonempty_row,
                "min_col": first_nonempty_col,
                "max_col": last_nonempty_col,
            },
            "freeze_panes": safe_value(ws.freeze_panes),
            "auto_filter": safe_value(ws.auto_filter.ref),
            "merged_ranges": [str(rng) for rng in ws.merged_cells.ranges],
            "formula_count": len(formulas),
            "formula_examples": formulas[:80],
            "wind_like_count": len(wind_like),
            "wind_like_examples": wind_like[:80],
            "chart_count": len(charts),
            "charts": charts,
            "image_count": len(getattr(ws, "_images", [])),
            "top_left_sample": {
                "range": f"{get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{max_row}",
                "values": cell_matrix(ws, min_row, max_row, min_col, max_col),
            },
            "label_examples": labels[:120],
        })

    report["formula_sheet_references"] = formula_ref_counter.most_common()
    report["formula_functions"] = formula_function_counter.most_common()
    output_path.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps({
        "sheets": len(report["sheets"]),
        "external_links": report["external_links"],
        "formula_functions": report["formula_functions"][:20],
        "output": str(output_path),
    }, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
