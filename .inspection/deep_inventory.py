"""
Step 1: Deep structural inventory of 9 target sheets.
Produces structured JSON for series_catalog, DATA_DICTIONARY, EXCEL_LINEAGE.
Read-only — never modifies the source workbook.
"""
import json
import sys
from pathlib import Path
from collections import defaultdict
import openpyxl
from openpyxl.utils import get_column_letter

WORKBOOK = Path("/Users/martin_ai/Desktop/Martin Monthly Brief/FX Chartbook - Flow 0515.xlsx")
TARGET_SHEETS = [
    "3.即远期", "3.代客即期", "3.涉外收付", "3.货物贸易",
    "3.贸易商", "3.服务贸易", "3.FDI", "3.证券EQ", "3.证券FI"
]
OUT_DIR = Path("/Users/martin_ai/Desktop/Martin Monthly Brief/.inspection")

def safe_value(cell):
    """Extract cell value safely."""
    if cell is None:
        return None
    if isinstance(cell.value, str):
        return cell.value.strip()
    return cell.value

def is_date_like(val):
    """Check if value looks like a date (int year-month or datetime)."""
    import datetime
    if isinstance(val, datetime.datetime):
        return True
    if isinstance(val, datetime.date):
        return True
    return False

def is_formula(cell):
    """Check if cell contains a formula."""
    if cell is None:
        return False
    return isinstance(cell.value, str) and cell.value.startswith("=")

def is_numeric(val):
    """Check if value is numeric."""
    return isinstance(val, (int, float)) and not isinstance(val, bool)

def scan_sheet(ws):
    """Deep scan a single worksheet."""
    result = {
        "name": ws.title,
        "dimensions": ws.dimensions,
        "max_row": ws.max_row,
        "max_column": ws.max_column,
        "freeze_panes": ws.freeze_panes,
        "merged_ranges": [str(m) for m in ws.merged_cells.ranges] if ws.merged_cells else [],
    }

    # Count cell types
    total_cells = 0
    formula_cells = 0
    numeric_cells = 0
    text_cells = 0
    empty_cells = 0
    date_cells = 0
    error_cells = 0
    zero_cells = 0

    # Track data regions
    rows_with_data = defaultdict(int)  # row -> nonempty count
    cols_with_data = defaultdict(int)  # col -> nonempty count

    # Sample first rows for layout analysis
    header_rows = []
    data_start_row = None

    for row_idx in range(1, min(ws.max_row + 1, 100)):  # Scan first 100 rows thoroughly
        row_data = []
        nonempty = 0
        for col_idx in range(1, min(ws.max_column + 1, 60)):
            cell = ws.cell(row=row_idx, column=col_idx)
            val = safe_value(cell)
            total_cells += 1

            if val is None:
                empty_cells += 1
            elif is_formula(cell):
                formula_cells += 1
                rows_with_data[row_idx] += 1
                cols_with_data[col_idx] += 1
            elif is_numeric(val):
                numeric_cells += 1
                rows_with_data[row_idx] += 1
                cols_with_data[col_idx] += 1
                if val == 0:
                    zero_cells += 1
            elif is_date_like(val):
                date_cells += 1
                rows_with_data[row_idx] += 1
                cols_with_data[col_idx] += 1
            elif isinstance(val, str):
                text_cells += 1
                rows_with_data[row_idx] += 1
                cols_with_data[col_idx] += 1
            else:
                text_cells += 1

            row_data.append({
                "col": get_column_letter(col_idx),
                "val": str(val)[:120] if val is not None else None,
                "is_formula": is_formula(cell),
            })

        if nonempty > 0 or row_idx <= 30:
            header_rows.append({
                "row": row_idx,
                "nonempty": rows_with_data[row_idx],
                "cells": row_data[:50]  # First 50 cols
            })

    # Find data regions (contiguous blocks with high data density)
    # A data region starts when row density jumps and continues while density stays high
    data_regions = []
    in_region = False
    region_start = None
    for row_idx in range(1, ws.max_row + 1):
        density = rows_with_data.get(row_idx, 0)
        if density >= 3 and not in_region:  # At least 3 nonempty cells to start a region
            region_start = row_idx
            in_region = True
        elif density < 2 and in_region:
            data_regions.append({"start_row": region_start, "end_row": row_idx - 1})
            in_region = False

    if in_region:
        data_regions.append({"start_row": region_start, "end_row": ws.max_row})

    # Column-level analysis
    column_headers = {}
    for col_idx in range(1, min(ws.max_column + 1, 60)):
        col_letter = get_column_letter(col_idx)
        # Try to find header in first 20 rows
        for row_idx in range(1, min(21, ws.max_row + 1)):
            cell = ws.cell(row=row_idx, column=col_idx)
            val = safe_value(cell)
            if isinstance(val, str) and len(val) > 0 and len(val) < 200:
                column_headers[col_letter] = {
                    "row": row_idx,
                    "header": val,
                    "nonempty_count": cols_with_data[col_idx],
                    "has_formulas": False,  # Check below
                }
                break

    # Check which columns have formulas in data rows
    for col_idx in range(1, min(ws.max_column + 1, 60)):
        col_letter = get_column_letter(col_idx)
        for row_idx in range(max(1, (data_regions[0]["start_row"] if data_regions else 20)), min(ws.max_row + 1, 101)):
            cell = ws.cell(row=row_idx, column=col_idx)
            if is_formula(cell):
                if col_letter in column_headers:
                    column_headers[col_letter]["has_formulas"] = True
                break

    # Find date columns (columns where first data rows are dates)
    date_columns = []
    for col_idx in range(1, min(ws.max_column + 1, 30)):
        date_count = 0
        total = 0
        for row_idx in range(1, min(ws.max_row + 1, 51)):
            cell = ws.cell(row=row_idx, column=col_idx)
            if is_date_like(safe_value(cell)):
                date_count += 1
            if safe_value(cell) is not None:
                total += 1
        if date_count > 5 and date_count / max(total, 1) > 0.5:
            date_columns.append(get_column_letter(col_idx))

    # Chart analysis
    charts = []
    if ws._charts:
        for i, ch in enumerate(ws._charts):
            try:
                anchor_str = str(ch.anchor) if hasattr(ch, 'anchor') else 'unknown'
            except Exception:
                anchor_str = 'repr_failed'
            charts.append({
                "index": i,
                "title": getattr(ch, 'title', ''),
                "type": str(type(ch).__name__),
                "anchor": anchor_str,
            })

    # Images
    images = []
    if ws._images:
        for i, img in enumerate(ws._images):
            try:
                anchor_str = str(img.anchor) if hasattr(img, 'anchor') else 'unknown'
            except Exception:
                anchor_str = 'repr_failed'
            images.append({
                "index": i,
                "anchor": anchor_str,
            })

    # Sample last rows (to find latest data)
    last_rows_sample = []
    for row_idx in range(max(1, ws.max_row - 5), ws.max_row + 1):
        row_cells = []
        for col_idx in range(1, min(ws.max_column + 1, 15)):
            cell = ws.cell(row=row_idx, column=col_idx)
            row_cells.append({
                "col": get_column_letter(col_idx),
                "val": str(safe_value(cell))[:80] if safe_value(cell) is not None else None,
            })
        last_rows_sample.append({"row": row_idx, "cells": row_cells})

    # External references (formulas with [1], external book references)
    ext_refs = set()
    for row_idx in range(1, min(ws.max_row + 1, 51)):
        for col_idx in range(1, min(ws.max_column + 1, 60)):
            cell = ws.cell(row=row_idx, column=col_idx)
            if is_formula(cell):
                ftext = str(cell.value)
                if '[' in ftext or '!' in ftext:
                    ext_refs.add(ftext[:200])

    # Key metrics
    result.update({
        "cell_stats": {
            "total_scanned": total_cells,
            "formula": formula_cells,
            "numeric": numeric_cells,
            "text": text_cells,
            "empty": empty_cells,
            "date": date_cells,
            "error": error_cells,
            "zero": zero_cells,
        },
        "data_regions": data_regions[:5],
        "column_headers": {k: v for k, v in list(column_headers.items())[:40]},
        "date_columns": date_columns,
        "charts": charts,
        "images": images,
        "last_rows_sample": last_rows_sample,
        "external_refs_sample": list(ext_refs)[:30],
        "header_rows_sample": header_rows[:35],
    })

    return result

def main():
    print(f"Opening {WORKBOOK} (read-only)...")
    wb = openpyxl.load_workbook(WORKBOOK, data_only=False, read_only=False)

    results = {}
    for sheet_name in TARGET_SHEETS:
        if sheet_name in wb.sheetnames:
            print(f"  Scanning: {sheet_name}")
            ws = wb[sheet_name]
            results[sheet_name] = scan_sheet(ws)
        else:
            print(f"  WARNING: {sheet_name} not found!")
            results[sheet_name] = {"error": "Sheet not found"}

    wb.close()

    # Save detailed report
    out_path = OUT_DIR / "deep_inventory.json"
    with open(out_path, 'w', encoding='utf-8') as f:
        json.dump(results, f, ensure_ascii=False, indent=2, default=str)
    print(f"\nSaved to {out_path}")
    print(f"Size: {out_path.stat().st_size} bytes")

    # Print summary table
    print("\n" + "="*100)
    print(f"{'Sheet':<20} {'Rows':>6} {'Cols':>6} {'Formulas':>10} {'Numeric':>10} {'Charts':>6} {'DateCols':<12}")
    print("-"*100)
    for name, r in results.items():
        if "error" in r:
            print(f"{name:<20} ERROR: {r['error']}")
        else:
            cs = r["cell_stats"]
            print(f"{name:<20} {r['max_row']:>6} {r['max_column']:>6} {cs['formula']:>10} {cs['numeric']:>10} {len(r['charts']):>6} {','.join(r['date_columns']):<12}")

if __name__ == "__main__":
    main()
