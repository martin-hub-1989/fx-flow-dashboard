"""
Step 1b: Extract actual data values (cached) and build comprehensive series catalog.
Reads with data_only=True to get cached values, and data_only=False for formulas.
"""
import json
import datetime
from pathlib import Path
from collections import defaultdict
import openpyxl

WORKBOOK = Path("/Users/martin_ai/Desktop/Martin Monthly Brief/FX Chartbook - Flow 0515.xlsx")
OUT_DIR = Path("/Users/martin_ai/Desktop/Martin Monthly Brief/config")
DOCS_DIR = Path("/Users/martin_ai/Desktop/Martin Monthly Brief/docs")

TARGET_SHEETS = [
    "3.即远期", "3.代客即期", "3.涉外收付", "3.货物贸易",
    "3.贸易商", "3.服务贸易", "3.FDI", "3.证券EQ", "3.证券FI"
]

def safe_val(cell):
    if cell is None or cell.value is None:
        return None
    if isinstance(cell.value, str):
        return cell.value.strip()
    return cell.value

def is_formula(cell):
    if cell is None:
        return False
    return isinstance(cell.value, str) and str(cell.value).startswith("=")

def detect_frequency(dates):
    """Detect frequency from a list of dates."""
    if len(dates) < 3:
        return "unknown"
    diffs = []
    for i in range(1, len(dates)):
        d = (dates[i] - dates[i-1]).days
        if d > 0:
            diffs.append(d)
    if not diffs:
        return "unknown"
    avg = sum(diffs) / len(diffs)
    if 25 <= avg <= 35:
        return "monthly"
    elif 80 <= avg <= 100:
        return "quarterly"
    elif 350 <= avg <= 380:
        return "annual"
    elif 1 <= avg <= 10:
        return "daily"
    elif avg < 1:
        return "daily"
    else:
        return "unknown"

def extract_sheet_series(sheet_name, ws_formulas, ws_values):
    """Extract all series from a sheet, combining formula and value info."""
    series_list = []

    # Determine max rows/cols
    max_row = ws_values.max_row
    max_col = ws_values.max_column

    # First, find the date column and data header row
    date_col = None
    date_header_row = None
    data_header_row = None

    # Scan first 30 rows for structure
    for row_idx in range(1, min(31, max_row + 1)):
        for col_idx in range(1, min(31, max_col + 1)):
            cell = ws_values.cell(row=row_idx, column=col_idx)
            val = safe_val(cell)
            if isinstance(val, (datetime.datetime, datetime.date)):
                if date_col is None:
                    date_col = col_idx
                    date_header_row = row_idx
                break

    # Find data start: row where date column becomes consistently populated
    date_count = 0
    for row_idx in range(1, min(101, max_row + 1)):
        cell = ws_values.cell(row=row_idx, column=date_col) if date_col else None
        val = safe_val(cell)
        if isinstance(val, (datetime.datetime, datetime.date)):
            date_count += 1
            if date_count == 1:
                data_header_row = row_idx

    # Scan column headers near the data header
    header_row = None
    if data_header_row:
        # Look for header text in rows near data start
        for row_idx in range(max(1, data_header_row - 3), min(data_header_row + 5, 30)):
            text_count = 0
            num_count = 0
            for col_idx in range(1, min(max_col + 1, 70)):
                cell = ws_values.cell(row=row_idx, column=col_idx)
                val = safe_val(cell)
                if isinstance(val, str) and len(val) > 1:
                    text_count += 1
                elif isinstance(val, (int, float)):
                    num_count += 1
            if text_count >= 2:
                header_row = row_idx
                break

    if header_row is None:
        header_row = data_header_row or 5

    # Extract dates
    dates = []
    date_values = []
    if date_col:
        for row_idx in range(data_header_row or 5, max_row + 1):
            cell = ws_values.cell(row=row_idx, column=date_col)
            val = safe_val(cell)
            if isinstance(val, datetime.datetime):
                dates.append(val.date())
                date_values.append(val)
            elif isinstance(val, datetime.date):
                dates.append(val)
                date_values.append(val)

    frequency = detect_frequency(dates)

    # Now extract each column as a potential series
    for col_idx in range(1, min(max_col + 1, 70)):
        # Get header
        header_cell = ws_values.cell(row=header_row, column=col_idx)
        header_val = safe_val(header_cell)

        # Check if this column has data
        data_count = 0
        formula_count = 0
        numeric_count = 0
        values_sample = []
        first_date = None
        last_date = None

        for row_idx in range(data_header_row or 5, max_row + 1):
            val_cell = ws_values.cell(row=row_idx, column=col_idx)
            val = safe_val(val_cell)
            f_cell = ws_formulas.cell(row=row_idx, column=col_idx)

            if val is not None:
                data_count += 1
                if isinstance(val, (int, float)):
                    numeric_count += 1
                    if len(values_sample) < 5:
                        if date_col and row_idx <= len(date_values):
                            d_idx = row_idx - (data_header_row or 5)
                            if 0 <= d_idx < len(date_values):
                                values_sample.append({"date": str(date_values[d_idx])[:10], "value": val})

                if first_date is None and date_col and row_idx <= len(dates):
                    d_idx = row_idx - (data_header_row or 5)
                    if 0 <= d_idx < len(dates):
                        first_date = str(dates[d_idx])

            if is_formula(f_cell):
                formula_count += 1

        # Find last date
        if dates:
            for row_idx in range(max_row, (data_header_row or 5) - 1, -1):
                val = safe_val(ws_values.cell(row=row_idx, column=col_idx))
                if val is not None and isinstance(val, (int, float)):
                    d_idx = row_idx - (data_header_row or 5)
                    if 0 <= d_idx < len(dates):
                        last_date = str(dates[d_idx])
                    break

        # Skip columns with almost no data
        if data_count < 5 and col_idx != date_col:
            continue

        # Determine series_type
        is_raw = formula_count == 0 and numeric_count > 0
        is_derived = formula_count > 0
        is_edb_ref = False

        # Check for EDB reference
        if header_val and "edb()" in str(header_val).lower():
            is_edb_ref = True

        # Build series entry
        col_letter = openpyxl.utils.get_column_letter(col_idx)
        series_id_base = f"{sheet_name.replace('3.','')}_{col_letter}"

        # Determine unit from header text
        unit = "unknown"
        if header_val:
            header_lower = str(header_val).lower()
            if "当月值" in str(header_val):
                unit = "monthly_value"
            elif "累计值" in str(header_val):
                unit = "cumulative_value"
            elif "累计同比" in str(header_val):
                unit = "cumulative_yoy_pct"
            elif "净流入" in str(header_val) or "净结汇" in str(header_val):
                unit = "net_flow"
            elif "占比" in str(header_val) or "比重" in str(header_val):
                unit = "percentage"
            elif "指数" in str(header_val):
                unit = "index"
            elif "收益率" in str(header_val):
                unit = "yield_pct"
            elif "差额" in str(header_val):
                unit = "balance"
            elif "mma" in header_lower or "mma" in header_lower:
                unit = "moving_average"
            elif "z-score" in header_lower or "z-score" in header_lower:
                unit = "z_score"
            elif "correl" in header_lower:
                unit = "correlation"
            elif "增速" in str(header_val) or "change" in header_lower:
                unit = "pct_change"

        entry = {
            "series_id": series_id_base,
            "display_name": str(header_val)[:200] if header_val else f"Column_{col_letter}",
            "module": sheet_name,
            "series_type": "derived" if is_derived else ("raw" if is_raw else "manual"),
            "frequency": frequency,
            "unit": unit,
            "source": "excel_seed",
            "excel_sheet": sheet_name,
            "excel_column": col_letter,
            "excel_header_row": header_row,
            "excel_data_start_row": data_header_row or 5,
            "first_date": first_date,
            "last_date": last_date,
            "observation_count": data_count,
            "formula_count": formula_count,
            "values_sample": values_sample,
            "has_edb_ref": is_edb_ref,
        }

        series_list.append(entry)

    return {
        "sheet": sheet_name,
        "max_row": max_row,
        "max_col": max_col,
        "date_column": openpyxl.utils.get_column_letter(date_col) if date_col else None,
        "date_header_row": date_header_row,
        "header_row": header_row,
        "frequency": frequency,
        "date_count": len(dates),
        "date_range": f"{dates[0]} to {dates[-1]}" if dates else "unknown",
        "series": series_list,
    }

def main():
    print("Opening workbook (formulas)...")
    wb_f = openpyxl.load_workbook(WORKBOOK, data_only=False, read_only=False)
    print("Opening workbook (values)...")
    wb_v = openpyxl.load_workbook(WORKBOOK, data_only=True, read_only=False)

    all_data = {}
    all_series_catalog = []

    for sheet_name in TARGET_SHEETS:
        if sheet_name not in wb_f.sheetnames:
            print(f"  SKIP: {sheet_name} not found")
            continue

        print(f"  Processing: {sheet_name}")
        ws_f = wb_f[sheet_name]
        ws_v = wb_v[sheet_name]

        result = extract_sheet_series(sheet_name, ws_f, ws_v)
        all_data[sheet_name] = result

        for s in result["series"]:
            # Create stable series_id
            module_prefix = {
                "3.即远期": "fx_fwd",
                "3.代客即期": "fx_cspot",
                "3.涉外收付": "fx_crossborder",
                "3.货物贸易": "trade_goods",
                "3.贸易商": "trade_merchant",
                "3.服务贸易": "trade_services",
                "3.FDI": "fdi",
                "3.证券EQ": "sec_eq",
                "3.证券FI": "sec_fi",
            }.get(sheet_name, sheet_name.replace("3.", ""))

            s["series_id"] = f"{module_prefix}:{s['excel_column']}"
            all_series_catalog.append(s)

    wb_f.close()
    wb_v.close()

    # Save full detail
    detail_path = OUT_DIR.parent / ".inspection" / "series_extract.json"
    detail_path.parent.mkdir(parents=True, exist_ok=True)
    with open(detail_path, 'w', encoding='utf-8') as f:
        json.dump(all_data, f, ensure_ascii=False, indent=2, default=str)

    # Save series catalog
    catalog_path = OUT_DIR / "series_catalog.json"
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    with open(catalog_path, 'w', encoding='utf-8') as f:
        json.dump(all_series_catalog, f, ensure_ascii=False, indent=2, default=str)

    # Print summary
    total_raw = sum(1 for s in all_series_catalog if s["series_type"] == "raw")
    total_derived = sum(1 for s in all_series_catalog if s["series_type"] == "derived")
    total_manual = sum(1 for s in all_series_catalog if s["series_type"] == "manual")

    print(f"\n{'='*80}")
    print(f"Series Catalog Summary")
    print(f"  Total series: {len(all_series_catalog)}")
    print(f"  Raw: {total_raw}, Derived: {total_derived}, Manual: {total_manual}")
    print(f"  Saved to: {catalog_path}")

    for sheet_name, data in all_data.items():
        raw_n = sum(1 for s in data["series"] if s["series_type"] == "raw")
        der_n = sum(1 for s in data["series"] if s["series_type"] == "derived")
        print(f"  {sheet_name}: {len(data['series'])} series ({raw_n} raw, {der_n} derived), freq={data['frequency']}, dates={data['date_range']}")

if __name__ == "__main__":
    main()
