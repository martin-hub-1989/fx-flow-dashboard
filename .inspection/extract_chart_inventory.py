import json
import re
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


WORKBOOK = Path("FX Chartbook - Flow 0515.xlsx")
TARGET_SHEETS = [
    "3.即远期",
    "3.代客即期",
    "3.涉外收付",
    "3.货物贸易",
    "3.贸易商",
    "3.服务贸易",
    "3.FDI",
    "3.证券EQ",
    "3.证券FI",
]


def rich_text(value):
    if value is None:
        return None
    try:
        parts = []
        for paragraph in value.tx.rich.p:
            for run in paragraph.r or []:
                if run.t:
                    parts.append(run.t)
            for field in paragraph.fld or []:
                if field.t:
                    parts.append(field.t)
        text = "".join(parts).strip()
        return text or None
    except Exception:
        return None


def series_title(series):
    tx = getattr(series, "tx", None)
    if tx is None:
        return None, None
    try:
        if tx.strRef and tx.strRef.f:
            return None, tx.strRef.f
    except Exception:
        pass
    try:
        if tx.v:
            return str(tx.v), None
    except Exception:
        pass
    return None, None


def ref_formula(obj):
    if obj is None:
        return None
    for ref_name in ("numRef", "strRef", "multiLvlStrRef"):
        ref = getattr(obj, ref_name, None)
        if ref is not None and getattr(ref, "f", None):
            return ref.f
    return None


def anchor_info(anchor):
    try:
        return {
            "from_cell": f"{get_column_letter(anchor._from.col + 1)}{anchor._from.row + 1}",
            "to_cell": f"{get_column_letter(anchor.to.col + 1)}{anchor.to.row + 1}",
            "width_cols": anchor.to.col - anchor._from.col,
            "height_rows": anchor.to.row - anchor._from.row,
        }
    except Exception:
        return {"repr": str(anchor)}


def chart_type_name(chart):
    parts = [type(chart).__name__]
    for subchart in getattr(chart, "_charts", []) or []:
        name = type(subchart).__name__
        if name not in parts:
            parts.append(name)
    return "+".join(parts)


def chart_series(chart):
    result = []
    all_charts = getattr(chart, "_charts", None) or [chart]
    seen = set()
    for subchart in all_charts:
        for idx, series in enumerate(getattr(subchart, "ser", []) or [], start=1):
            title_value, title_ref = series_title(series)
            values_ref = ref_formula(getattr(series, "val", None))
            if values_ref is None:
                values_ref = ref_formula(getattr(series, "yVal", None))
            categories_ref = ref_formula(getattr(series, "cat", None))
            if categories_ref is None:
                categories_ref = ref_formula(getattr(series, "xVal", None))
            key = (title_value, title_ref, values_ref, categories_ref)
            if key in seen:
                continue
            seen.add(key)
            result.append({
                "index": idx,
                "chart_type": type(subchart).__name__,
                "title_value": title_value,
                "title_ref": title_ref,
                "values_ref": values_ref,
                "categories_ref": categories_ref,
                "axis_group": getattr(series, "axId", None),
            })
    return result


def resolve_cell_ref(workbook, formula):
    if not formula:
        return None
    match = re.match(
        r"^'?(?P<sheet>[^']+?)'?!\$(?P<col>[A-Z]+)\$(?P<row>\d+)$",
        formula,
    )
    if not match:
        return None
    sheet_name = match.group("sheet")
    if sheet_name not in workbook.sheetnames:
        return None
    return workbook[sheet_name][f"{match.group('col')}{match.group('row')}"].value


def main():
    workbook = load_workbook(WORKBOOK, data_only=False, read_only=False)
    inventory = []
    for sheet_name in TARGET_SHEETS:
        sheet = workbook[sheet_name]
        for index, chart in enumerate(sheet._charts, start=1):
            title = rich_text(chart.title)
            series = chart_series(chart)
            for item in series:
                item["resolved_title"] = (
                    item["title_value"]
                    or resolve_cell_ref(workbook, item["title_ref"])
                )
            inventory.append({
                "chart_id": f"{sheet_name}#{index:02d}",
                "sheet": sheet_name,
                "index": index,
                "type": chart_type_name(chart),
                "title": title,
                "style": getattr(chart, "style", None),
                "grouping": getattr(chart, "grouping", None),
                "overlap": getattr(chart, "overlap", None),
                "anchor": anchor_info(chart.anchor),
                "series": series,
            })

    output = Path(".inspection/chart_inventory.json")
    output.write_text(
        json.dumps(inventory, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    print(f"charts={len(inventory)} output={output}")
    for chart in inventory:
        refs = [
            item["resolved_title"]
            or item["title_ref"]
            or item["title_value"]
            or item["values_ref"]
            for item in chart["series"]
        ]
        print(
            f"{chart['chart_id']}\t{chart['type']}\t"
            f"{chart['title'] or '(untitled)'}\t{len(chart['series'])}\t"
            f"{' | '.join(str(ref) for ref in refs[:4])}"
        )


if __name__ == "__main__":
    main()
