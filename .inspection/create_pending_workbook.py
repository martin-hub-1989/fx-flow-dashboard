import shutil
import sys
from pathlib import Path

from openpyxl import load_workbook


CORE_SHEETS = [
    "3.即远期",
    "3.代客即期",
    "3.涉外收付",
    "3.货物贸易",
    "3.贸易商",
    "3.服务贸易",
    "3.FDI",
    "3.证券EQ",
    "3.证券FI",
]


def main():
    source = Path(sys.argv[1])
    output = Path(sys.argv[2])
    shutil.copy2(source, output)

    workbook = load_workbook(output, data_only=False, read_only=False)
    missing = [name for name in CORE_SHEETS if name not in workbook.sheetnames]
    if missing:
        raise RuntimeError(f"Missing core sheets: {missing}")

    for name in CORE_SHEETS:
        workbook.remove(workbook[name])

    workbook.save(output)
    print("\n".join(workbook.sheetnames))


if __name__ == "__main__":
    main()
